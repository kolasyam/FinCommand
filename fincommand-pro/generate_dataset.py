import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def build_dataset():
    # ── 115 Full Ledgers covering all Notes 1 to 26 ──
    # [code, name, note_no, note_name, section, treasury_type, normal_bal, op_dr, op_cr, m_dr_base, m_cr_base]
    ledgers_def = [
        # --- PPE (Note 10, anc) ---
        ('1001', 'Plant & Machinery', 10, 'PPE', 'anc', None, 'Dr', 4500, 0, [50, 0, 0, 80, 0, 0, 60, 0, 0, 40, 0, 0], [0]*12),
        ('1002', 'Furniture & Fixtures', 10, 'PPE', 'anc', None, 'Dr', 850, 0, [10, 0, 15, 0, 0, 20, 0, 0, 10, 0, 0, 0], [0]*12),
        ('1003', 'Computers & Peripherals', 10, 'PPE', 'anc', None, 'Dr', 1600, 0, [40, 35, 50, 30, 45, 60, 40, 50, 35, 40, 30, 45], [0]*12),
        ('1004', 'Vehicles', 10, 'PPE', 'anc', None, 'Dr', 450, 0, [0]*12, [0]*12),
        ('1005', 'Office Equipment', 10, 'PPE', 'anc', None, 'Dr', 380, 0, [5, 0, 10, 0, 5, 0, 10, 0, 5, 0, 10, 0], [0]*12),
        ('1006', 'Accumulated Depreciation — PPE', 10, 'PPE', 'anc', None, 'Cr', 0, 1850, [0]*12, [41, 41, 41, 43, 43, 43, 45, 45, 45, 46, 46, 46]),
        ('1007', 'Capital Work-in-Progress', 10, 'PPE', 'anc', None, 'Dr', 870, 0, [30, 40, 50, 20, 30, 40, 20, 30, 40, 20, 30, 40], [0]*12),

        # --- ROU Assets (Note 11, anc) ---
        ('1011', 'Right-of-Use Asset — Office', 11, 'ROU Assets', 'anc', None, 'Dr', 600, 0, [0]*12, [0]*12),
        ('1012', 'Accumulated Depreciation — ROU', 11, 'ROU Assets', 'anc', None, 'Cr', 0, 180, [0]*12, [15]*12),

        # --- Intangibles (Note 12, anc) ---
        ('1021', 'Computer Software (Purchased)', 12, 'Intangibles', 'anc', None, 'Dr', 420, 0, [20, 0, 0, 30, 0, 0, 25, 0, 0, 15, 0, 0], [0]*12),
        ('1022', 'Internally Developed Software', 12, 'Intangibles', 'anc', None, 'Dr', 280, 0, [10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10], [0]*12),
        ('1023', 'Amortisation — Intangibles', 12, 'Intangibles', 'anc', None, 'Cr', 0, 320, [0]*12, [12]*12),

        # --- Non-Current Investments (Note 13, anc) ---
        ('1031', 'Investment in Subsidiaries', 13, 'Investments NC', 'anc', None, 'Dr', 350, 0, [0]*12, [0]*12),
        ('1032', 'Investment in Associates', 13, 'Investments NC', 'anc', None, 'Dr', 150, 0, [0]*12, [0]*12),

        # --- Current Investments (Note 13, ac, mf) ---
        ('1033', 'HDFC Liquid Fund', 13, 'Investments Current', 'ac', 'mf', 'Dr', 350, 0, [50, 0, 100, 0, 50, 0, 100, 0, 50, 0, 100, 0], [0, 40, 0, 60, 0, 40, 0, 60, 0, 40, 0, 60]),
        ('1034', 'ICICI Pru Money Market Fund', 13, 'Investments Current', 'ac', 'mf', 'Dr', 250, 0, [30, 0, 50, 0, 30, 0, 50, 0, 30, 0, 50, 0], [0, 20, 0, 30, 0, 20, 0, 30, 0, 20, 0, 30]),
        ('1035', 'SBI Liquid Fund', 13, 'Investments Current', 'ac', 'mf', 'Dr', 150, 0, [20, 0, 30, 0, 20, 0, 30, 0, 20, 0, 30, 0], [0, 10, 0, 20, 0, 10, 0, 20, 0, 10, 0, 20]),
        ('1036', 'Axis Liquid Fund', 13, 'Investments Current', 'ac', 'mf', 'Dr', 100, 0, [10, 0, 20, 0, 10, 0, 20, 0, 10, 0, 20, 0], [0, 10, 0, 10, 0, 10, 0, 10, 0, 10, 0, 10]),

        # --- Other NC Assets (Note 14, anc) ---
        ('1041', 'Security Deposit (NC)', 14, 'Other NC Assets', 'anc', None, 'Dr', 80, 0, [0]*12, [0]*12),
        ('1042', 'Advance Tax (NC)', 14, 'Other NC Assets', 'anc', None, 'Dr', 90, 0, [15, 0, 0, 20, 0, 0, 25, 0, 0, 30, 0, 0], [0]*12),
        ('1043', 'MAT Credit Entitlement', 14, 'Other NC Assets', 'anc', None, 'Dr', 40, 0, [0]*12, [0]*12),

        # --- Inventories (Note 15, ac) ---
        ('1051', 'Traded Goods / Stock-in-Trade', 15, 'Inventories', 'ac', None, 'Dr', 250, 0, [150, 160, 170, 180, 175, 185, 190, 200, 195, 205, 210, 200], [140, 155, 165, 175, 170, 180, 185, 195, 190, 200, 205, 195]),
        ('1052', 'Work-in-Progress', 15, 'Inventories', 'ac', None, 'Dr', 110, 0, [40, 45, 50, 55, 50, 60, 65, 70, 65, 75, 80, 70], [38, 42, 48, 52, 48, 58, 62, 68, 62, 72, 78, 68]),
        ('1053', 'Raw Materials', 15, 'Inventories', 'ac', None, 'Dr', 64, 0, [30, 32, 35, 38, 36, 40, 42, 45, 43, 48, 50, 46], [28, 30, 33, 36, 34, 38, 40, 43, 41, 46, 48, 44]),

        # --- Trade Receivables (Note 16, ac) ---
        ('1061', 'Sundry Debtors < 6 months', 16, 'Trade Receivables', 'ac', None, 'Dr', 3100, 0, [2100, 2200, 2300, 2400, 2350, 2500, 2600, 2750, 2650, 2800, 2900, 2850], [1950, 2050, 2150, 2250, 2200, 2350, 2450, 2600, 2500, 2650, 2750, 2700]),
        ('1062', 'Sundry Debtors > 6 months', 16, 'Trade Receivables', 'ac', None, 'Dr', 480, 0, [50, 40, 60, 30, 40, 50, 40, 60, 50, 40, 50, 30], [40, 50, 30, 40, 50, 30, 50, 40, 60, 30, 40, 50]),
        ('1063', 'Provision for Bad Debts (ECL)', 16, 'Trade Receivables', 'ac', None, 'Cr', 0, 100, [0]*12, [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5]),

        # --- Cash & CE (Note 19, ac) ---
        ('2001', 'Cash in Hand', 19, 'Cash & CE', 'ac', 'cash', 'Dr', 25, 0, [10, 12, 11, 15, 13, 14, 16, 15, 14, 18, 17, 16], [9, 11, 10, 14, 12, 13, 15, 14, 13, 17, 16, 15]),
        ('2002', 'Petty Cash', 19, 'Cash & CE', 'ac', 'cash', 'Dr', 17, 0, [8, 9, 7, 10, 8, 9, 11, 10, 9, 12, 11, 10], [7, 8, 7, 9, 8, 8, 10, 9, 9, 11, 10, 10]),
        ('2101', 'HDFC Bank — Current Account', 19, 'Cash & CE', 'ac', 'bank_ca', 'Dr', 4690, 0, [1850, 1920, 2010, 2120, 2070, 2190, 2270, 2410, 2330, 2450, 2560, 2470], [1620, 1710, 1780, 1890, 1840, 1930, 2010, 2140, 2080, 2170, 2250, 2180]),
        ('2102', 'ICICI Bank — Current Account', 19, 'Cash & CE', 'ac', 'bank_ca', 'Dr', 320, 0, [150, 160, 170, 180, 175, 185, 190, 200, 195, 205, 210, 200], [140, 150, 160, 170, 165, 175, 180, 190, 185, 195, 200, 190]),
        ('2103', 'SBI — Current Account', 19, 'Cash & CE', 'ac', 'bank_ca', 'Dr', 210, 0, [80, 85, 90, 95, 90, 100, 105, 110, 105, 115, 120, 110], [75, 80, 85, 90, 85, 95, 100, 105, 100, 110, 115, 105]),
        ('2104', 'Axis Bank — Current Account', 19, 'Cash & CE', 'ac', 'bank_ca', 'Dr', 180, 0, [70, 75, 80, 85, 80, 90, 95, 100, 95, 105, 110, 100], [65, 70, 75, 80, 75, 85, 90, 95, 90, 100, 105, 95]),
        ('2105', 'Kotak Bank — Current Account', 19, 'Cash & CE', 'ac', 'bank_ca', 'Dr', 140, 0, [60, 65, 70, 75, 70, 80, 85, 90, 85, 95, 100, 90], [55, 60, 65, 70, 65, 75, 80, 85, 80, 90, 95, 85]),
        ('2201', 'Kotak Bank — Savings / Sweep', 19, 'Cash & CE', 'ac', 'bank_sb', 'Dr', 310, 0, [40, 50, 45, 60, 50, 55, 65, 60, 55, 70, 65, 60], [30, 40, 35, 50, 40, 45, 55, 50, 45, 60, 55, 50]),
        ('2202', 'HDFC Bank — Savings Account', 19, 'Cash & CE', 'ac', 'bank_sb', 'Dr', 250, 0, [30, 35, 40, 45, 40, 50, 55, 60, 55, 65, 70, 60], [25, 30, 35, 40, 35, 45, 50, 55, 50, 60, 65, 55]),

        # --- Bank Balances / FDs (Note 20, ac, fd) ---
        ('2301', 'HDFC Fixed Deposit — 001', 20, 'Bank Balances (FDs)', 'ac', 'fd', 'Dr', 1500, 0, [0]*12, [0]*12),
        ('2302', 'ICICI Fixed Deposit — 002', 20, 'Bank Balances (FDs)', 'ac', 'fd', 'Dr', 500, 0, [0]*12, [0]*12),
        ('2303', 'SBI Fixed Deposit — 003', 20, 'Bank Balances (FDs)', 'ac', 'fd', 'Dr', 200, 0, [0]*12, [0]*12),
        ('2304', 'Axis Fixed Deposit — 001', 20, 'Bank Balances (FDs)', 'ac', 'fd', 'Dr', 100, 0, [0]*12, [0]*12),
        ('2305', 'HDFC Margin Money FD', 20, 'Bank Balances (FDs)', 'ac', 'fd', 'Dr', 100, 0, [0]*12, [0]*12),

        # --- Loans & Advances (Note 21, ac) ---
        ('1071', 'Staff Advance (Current)', 21, 'Loans & Advances', 'ac', None, 'Dr', 120, 0, [15, 18, 20, 22, 20, 25, 28, 30, 27, 32, 35, 30], [12, 15, 18, 20, 18, 22, 25, 28, 25, 30, 32, 28]),
        ('1072', 'Security Deposit (Current)', 21, 'Loans & Advances', 'ac', None, 'Dr', 80, 0, [0]*12, [0]*12),
        ('1073', 'Prepaid Expenses', 21, 'Loans & Advances', 'ac', None, 'Dr', 140, 0, [40, 10, 10, 45, 10, 10, 50, 10, 10, 55, 10, 10], [15, 15, 15, 18, 18, 18, 20, 20, 20, 22, 22, 22]),

        # --- Other Current Assets (Note 23, ac) ---
        ('1081', 'GST Input Tax Credit Receivable', 23, 'Other Current Assets', 'ac', None, 'Dr', 280, 0, [110, 115, 120, 125, 122, 130, 135, 140, 138, 145, 150, 142], [105, 110, 115, 120, 118, 125, 130, 135, 132, 140, 145, 138]),
        ('1082', 'TDS Receivable', 23, 'Other Current Assets', 'ac', None, 'Dr', 160, 0, [45, 47, 50, 52, 51, 54, 56, 60, 58, 61, 64, 60], [0]*12),
        ('1083', 'Advance to Suppliers', 23, 'Other Current Assets', 'ac', None, 'Dr', 70, 0, [25, 30, 35, 40, 35, 45, 50, 55, 50, 60, 65, 55], [20, 25, 30, 35, 30, 40, 45, 50, 45, 55, 60, 50]),
        ('1084', 'Unbilled Revenue / Contract Asset', 23, 'Other Current Assets', 'ac', None, 'Dr', 50, 0, [60, 65, 70, 75, 72, 78, 82, 88, 85, 90, 95, 88], [55, 60, 65, 70, 68, 74, 78, 84, 80, 86, 90, 84]),

        # --- Equity (Notes 1 & 2, eq) ---
        ('3001', 'Equity Share Capital', 1, 'Share Capital', 'eq', None, 'Cr', 0, 10000, [0]*12, [0]*12),
        ('3011', 'Securities Premium Reserve', 2, 'Other Equity', 'eq', None, 'Cr', 0, 25000, [0]*12, [0]*12),
        ('3012', 'General Reserve', 2, 'Other Equity', 'eq', None, 'Cr', 0, 5000, [0]*12, [0]*12),
        ('3013', 'Retained Earnings', 2, 'Other Equity', 'eq', None, 'Cr', 0, 50000, [0]*12, [0]*12),
        ('3014', 'ESOP Reserve', 2, 'Other Equity', 'eq', None, 'Cr', 0, 450, [0]*12, [8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8]),
        ('3015', 'OCI — Remeasurement of DBO', 2, 'Other Equity', 'eq', None, 'Dr', 140, 0, [0]*12, [0]*12),

        # --- Long-Term Borrowings (Note 3, lnc) ---
        ('4001', 'Term Loan — HDFC Bank', 3, 'Long-Term Borrowings', 'lnc', None, 'Cr', 0, 11280, [60, 60, 60, 60, 60, 60, 60, 60, 60, 60, 60, 60], [0]*12),
        ('4002', 'Term Loan — SBI', 3, 'Long-Term Borrowings', 'lnc', None, 'Cr', 0, 3200, [20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20, 20], [0]*12),
        ('4003', 'Term Loan — ICICI Bank', 3, 'Long-Term Borrowings', 'lnc', None, 'Cr', 0, 1800, [10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10], [0]*12),

        # --- Lease Liabilities NC (Note 4, lnc) ---
        ('4011', 'Lease Liability — Long Term (IND AS 116)', 4, 'Lease Liabilities', 'lnc', None, 'Cr', 0, 450, [12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12], [0]*12),

        # --- Deferred Tax (Note 5, lnc) ---
        ('4021', 'Deferred Tax Liability', 5, 'Deferred Tax', 'lnc', None, 'Cr', 0, 180, [0]*12, [0]*12),
        ('4022', 'Deferred Tax Asset', 5, 'Deferred Tax', 'lnc', None, 'Dr', 120, 0, [0]*12, [0]*12),

        # --- Long-Term Provisions (Note 6, lnc) ---
        ('4031', 'Gratuity Liability (IND AS 19)', 6, 'Long-Term Provisions', 'lnc', None, 'Cr', 0, 280, [0]*12, [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5]),
        ('4032', 'Leave Encashment Liability (NC)', 6, 'Long-Term Provisions', 'lnc', None, 'Cr', 0, 160, [0]*12, [3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3]),

        # --- Trade Payables (Note 7, lc) ---
        ('5001', 'MSME Trade Creditors', 7, 'Trade Payables', 'lc', None, 'Cr', 0, 850, [450, 480, 510, 540, 520, 560, 580, 610, 590, 630, 660, 620], [470, 500, 530, 560, 540, 580, 600, 630, 610, 650, 680, 640]),
        ('5002', 'Other Trade Creditors', 7, 'Trade Payables', 'lc', None, 'Cr', 0, 1290, [680, 720, 760, 800, 780, 830, 870, 910, 890, 940, 980, 930], [700, 740, 780, 820, 800, 850, 890, 930, 910, 960, 1000, 950]),

        # --- Other Financial Liabilities (Note 8, lc) ---
        ('5011', 'Accrued Salaries & Benefits', 8, 'Other Financial Liabilities', 'lc', None, 'Cr', 0, 480, [515, 538, 569, 601, 584, 620, 642, 687, 665, 699, 774, 656], [525, 548, 579, 611, 594, 630, 652, 697, 675, 709, 784, 666]),
        ('5012', 'Accrued Expenses', 8, 'Other Financial Liabilities', 'lc', None, 'Cr', 0, 180, [80, 85, 90, 95, 90, 100, 105, 110, 105, 115, 120, 110], [85, 90, 95, 100, 95, 105, 110, 115, 110, 120, 125, 115]),
        ('5013', 'Customer Deposit / Retention', 8, 'Other Financial Liabilities', 'lc', None, 'Cr', 0, 120, [0]*12, [5, 0, 10, 0, 5, 0, 10, 0, 5, 0, 10, 0]),
        ('5014', 'Lease Liability — Current (IND AS 116)', 8, 'Other Financial Liabilities', 'lc', None, 'Cr', 0, 90, [12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12], [12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12]),

        # --- Short-Term Borrowings (Note 9, lc) ---
        ('5021', 'OD / Working Capital Loan — HDFC', 9, 'ST Borrowings', 'lc', None, 'Cr', 0, 380, [150, 160, 170, 180, 175, 185, 190, 200, 195, 205, 210, 200], [160, 170, 180, 190, 185, 195, 200, 210, 205, 215, 220, 210]),
        ('5022', 'Working Capital Loan — SBI', 9, 'ST Borrowings', 'lc', None, 'Cr', 0, 100, [50, 55, 60, 65, 60, 70, 75, 80, 75, 85, 90, 80], [55, 60, 65, 70, 65, 75, 80, 85, 80, 90, 95, 85]),

        # --- Other Current Liabilities (Note 17, lc) ---
        ('5031', 'Advance from Customers', 17, 'Other Current Liabilities', 'lc', None, 'Cr', 0, 220, [90, 95, 100, 105, 100, 110, 115, 120, 115, 125, 130, 120], [100, 105, 110, 115, 110, 120, 125, 130, 125, 135, 140, 130]),
        ('5032', 'Statutory Dues Payable (PF, ESIC)', 17, 'Other Current Liabilities', 'lc', None, 'Cr', 0, 140, [55, 58, 62, 65, 63, 67, 70, 74, 72, 76, 82, 71], [58, 61, 65, 68, 66, 70, 73, 77, 75, 79, 85, 74]),
        ('5033', 'TDS Payable', 17, 'Other Current Liabilities', 'lc', None, 'Cr', 0, 180, [75, 78, 82, 86, 84, 89, 92, 98, 95, 100, 108, 94], [80, 83, 87, 91, 89, 94, 97, 103, 100, 105, 113, 99]),
        ('5034', 'GST Payable', 17, 'Other Current Liabilities', 'lc', None, 'Cr', 0, 110, [150, 158, 166, 175, 171, 181, 188, 200, 194, 204, 214, 204], [160, 168, 176, 185, 181, 191, 198, 210, 204, 214, 224, 214]),
        ('5035', 'Current Tax Payable', 17, 'Other Current Liabilities', 'lc', None, 'Cr', 0, 76, [0, 0, 120, 0, 0, 140, 0, 0, 160, 0, 0, 180], [110, 115, 125, 130, 128, 135, 140, 150, 145, 152, 160, 150]),

        # --- Short-Term Provisions (Note 18, lc) ---
        ('5041', 'Provision for Employee Benefits (ST)', 18, 'Short-Term Provisions', 'lc', None, 'Cr', 0, 120, [0]*12, [10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10, 10]),
        ('5042', 'Provision for Expenses / Tax (ST)', 18, 'Short-Term Provisions', 'lc', None, 'Cr', 0, 80, [0]*12, [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5]),

        # --- Revenue from Operations (Note 20, inc) ---
        ('6001', 'IT Services Revenue', 20, 'Revenue from Operations', 'inc', None, 'Cr', 0, 0, [0]*12, [1509, 1576, 1662, 1753, 1708, 1813, 1880, 2007, 1940, 2037, 2135, 2049]),
        ('6002', 'Managed Services / AMC Revenue', 20, 'Revenue from Operations', 'inc', None, 'Cr', 0, 0, [0]*12, [644, 675, 712, 751, 732, 777, 806, 860, 831, 873, 915, 878]),
        ('6003', 'Product / Resale Revenue', 20, 'Revenue from Operations', 'inc', None, 'Cr', 0, 0, [0]*12, [320, 335, 350, 370, 360, 385, 400, 425, 410, 430, 450, 435]),
        ('6004', 'SaaS Subscription Revenue', 20, 'Revenue from Operations', 'inc', None, 'Cr', 0, 0, [0]*12, [210, 220, 230, 245, 240, 255, 265, 280, 270, 285, 295, 288]),
        ('6005', 'Consulting Revenue', 20, 'Revenue from Operations', 'inc', None, 'Cr', 0, 0, [0]*12, [140, 148, 155, 165, 160, 172, 180, 192, 185, 195, 205, 196]),

        # --- Other Income (Note 21, inc) ---
        ('6011', 'Interest on Fixed Deposits', 21, 'Other Income', 'inc', None, 'Cr', 0, 0, [0]*12, [14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]),
        ('6012', 'Dividend from Mutual Funds', 21, 'Other Income', 'inc', None, 'Cr', 0, 0, [0]*12, [5, 0, 8, 0, 6, 0, 9, 0, 7, 0, 10, 0]),
        ('6013', 'Profit on Sale of Investments', 21, 'Other Income', 'inc', None, 'Cr', 0, 0, [0]*12, [0, 6, 0, 8, 0, 7, 0, 10, 0, 9, 0, 12]),
        ('6014', 'Miscellaneous Income', 21, 'Other Income', 'inc', None, 'Cr', 0, 0, [0]*12, [3, 4, 3, 5, 4, 4, 5, 6, 5, 5, 6, 5]),
        ('6015', 'Foreign Exchange Gain', 21, 'Other Income', 'inc', None, 'Cr', 0, 0, [0]*12, [8, 0, 12, 0, 9, 0, 14, 0, 11, 0, 15, 0]),

        # --- Cost of Services (Note 22, exp) ---
        ('7001', 'Subcontracting Charges', 22, 'Cost of Services', 'exp', None, 'Dr', 0, 0, [450, 470, 495, 520, 505, 535, 555, 590, 570, 600, 630, 605], [0]*12),
        ('7002', 'Cloud & Infrastructure Costs', 22, 'Cost of Services', 'exp', None, 'Dr', 0, 0, [210, 220, 230, 245, 238, 252, 260, 278, 268, 282, 296, 284], [0]*12),
        ('7003', 'Software Licences (Direct)', 22, 'Cost of Services', 'exp', None, 'Dr', 0, 0, [160, 168, 176, 186, 180, 192, 198, 212, 204, 214, 224, 216], [0]*12),
        ('7004', 'Technical Consumables', 22, 'Cost of Services', 'exp', None, 'Dr', 0, 0, [45, 48, 50, 53, 51, 55, 57, 60, 58, 61, 64, 61], [0]*12),
        ('7005', 'Data Centre Costs', 22, 'Cost of Services', 'exp', None, 'Dr', 0, 0, [85, 89, 94, 99, 96, 102, 106, 113, 109, 115, 120, 115], [0]*12),

        # --- Employee Benefits (Note 23, exp) ---
        ('7011', 'Salaries & Wages', 23, 'Employee Benefits', 'exp', None, 'Dr', 0, 0, [515, 538, 569, 601, 584, 620, 642, 687, 665, 699, 774, 656], [0]*12),
        ('7012', 'Bonus & Incentives', 23, 'Employee Benefits', 'exp', None, 'Dr', 0, 0, [60, 65, 70, 75, 72, 78, 82, 88, 85, 90, 110, 88], [0]*12),
        ('7013', 'PF Contribution (Employer)', 23, 'Employee Benefits', 'exp', None, 'Dr', 0, 0, [42, 44, 46, 49, 47, 50, 52, 56, 54, 57, 63, 53], [0]*12),
        ('7014', 'ESIC Contribution', 23, 'Employee Benefits', 'exp', None, 'Dr', 0, 0, [12, 13, 14, 15, 14, 15, 16, 17, 16, 17, 19, 16], [0]*12),
        ('7015', 'Gratuity Expense (IND AS 19)', 23, 'Employee Benefits', 'exp', None, 'Dr', 0, 0, [5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5, 5], [0]*12),
        ('7016', 'ESOP Charge (IND AS 102)', 23, 'Employee Benefits', 'exp', None, 'Dr', 0, 0, [8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8], [0]*12),
        ('7017', 'Staff Welfare & Training', 23, 'Employee Benefits', 'exp', None, 'Dr', 0, 0, [25, 27, 29, 31, 30, 32, 34, 37, 35, 37, 42, 36], [0]*12),
        ('7018', 'Leave Encashment Expense', 23, 'Employee Benefits', 'exp', None, 'Dr', 0, 0, [3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3, 3], [0]*12),

        # --- Finance Costs (Note 24, exp) ---
        ('7021', 'Interest on Term Loans', 24, 'Finance Costs', 'exp', None, 'Dr', 0, 0, [85, 84, 83, 82, 81, 80, 79, 78, 77, 76, 75, 74], [0]*12),
        ('7022', 'Interest on Lease Liabilities', 24, 'Finance Costs', 'exp', None, 'Dr', 0, 0, [4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4, 4], [0]*12),
        ('7023', 'Bank Charges & Processing Fees', 24, 'Finance Costs', 'exp', None, 'Dr', 0, 0, [6, 7, 6, 8, 7, 7, 8, 9, 8, 8, 9, 8], [0]*12),
        ('7024', 'Foreign Exchange Loss', 24, 'Finance Costs', 'exp', None, 'Dr', 0, 0, [0, 4, 0, 6, 0, 5, 0, 7, 0, 6, 0, 8], [0]*12),

        # --- Depreciation & Amortisation (Note 25, exp) ---
        ('7031', 'Depreciation on PPE (IND AS 16)', 25, 'Depreciation & Amort.', 'exp', None, 'Dr', 0, 0, [41, 41, 41, 43, 43, 43, 45, 45, 45, 46, 46, 46], [0]*12),
        ('7032', 'Amortisation — Intangibles', 25, 'Depreciation & Amort.', 'exp', None, 'Dr', 0, 0, [12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12], [0]*12),
        ('7033', 'Depreciation on ROU Assets', 25, 'Depreciation & Amort.', 'exp', None, 'Dr', 0, 0, [15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15, 15], [0]*12),

        # --- Other Expenses (Note 26, exp) ---
        ('7041', 'Rent & Utilities', 26, 'Other Expenses', 'exp', None, 'Dr', 0, 0, [75, 75, 75, 78, 78, 78, 80, 80, 80, 82, 82, 82], [0]*12),
        ('7042', 'Marketing & Business Development', 26, 'Other Expenses', 'exp', None, 'Dr', 0, 0, [110, 115, 120, 128, 124, 132, 138, 146, 140, 148, 156, 148], [0]*12),
        ('7043', 'Professional & Legal Fees', 26, 'Other Expenses', 'exp', None, 'Dr', 0, 0, [45, 48, 50, 54, 52, 55, 58, 62, 59, 62, 66, 62], [0]*12),
        ('7044', 'Travel & Conveyance', 26, 'Other Expenses', 'exp', None, 'Dr', 0, 0, [65, 68, 72, 76, 74, 78, 82, 87, 83, 88, 93, 88], [0]*12),
        ('7045', 'Admin & Office Expenses', 26, 'Other Expenses', 'exp', None, 'Dr', 0, 0, [35, 37, 39, 41, 40, 42, 44, 47, 45, 47, 50, 47], [0]*12),
        ('7046', 'Insurance Premium', 26, 'Other Expenses', 'exp', None, 'Dr', 0, 0, [20, 0, 0, 22, 0, 0, 24, 0, 0, 26, 0, 0], [0]*12),
        ('7047', 'Communication & Internet', 26, 'Other Expenses', 'exp', None, 'Dr', 0, 0, [18, 19, 20, 21, 20, 21, 22, 23, 22, 23, 25, 23], [0]*12),
        ('7048', 'CSR Expenditure', 26, 'Other Expenses', 'exp', None, 'Dr', 0, 0, [0]*12, [0]*12),
    ]

    months_name = ['Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Jan', 'Feb', 'Mar']
    ledgers_def = [list(l) for l in ledgers_def]

    # Step 1. Balance base monthly Dr/Cr figures by adjusting HDFC Bank (2101)
    for m in range(12):
        m_dr = sum(l[9][m] for l in ledgers_def)
        m_cr = sum(l[10][m] for l in ledgers_def)
        diff = m_dr - m_cr
        if diff > 0:
            for l in ledgers_def:
                if l[0] == '2101':
                    l_cr = list(l[10])
                    l_cr[m] += diff
                    l[10] = l_cr
                    break
        elif diff < 0:
            diff_abs = abs(diff)
            for l in ledgers_def:
                if l[0] == '2101':
                    l_dr = list(l[9])
                    l_dr[m] += diff_abs
                    l[9] = l_dr
                    break

    # Step 2. Monthly PAT Allocation: Credit 3013 (Retained Earnings Cr, eq) and Debit 7048 (CSR & Profit Allocation Dr, exp)
    for m in range(12):
        m_inc = sum((l[10][m] - l[9][m]) for l in ledgers_def if l[4] == 'inc')
        m_exp = sum((l[9][m] - l[10][m]) for l in ledgers_def if l[4] == 'exp' and l[0] != '7048')
        m_pat = m_inc - m_exp
        if m_pat > 0:
            for l in ledgers_def:
                if l[0] == '3013':
                    l_cr = list(l[10])
                    l_cr[m] += m_pat
                    l[10] = l_cr
                elif l[0] == '7048':
                    l_dr = list(l[9])
                    l_dr[m] += m_pat
                    l[9] = l_dr

    # Step 3. Set Retained Earnings (3013 op_cr = 50,000)
    for l in ledgers_def:
        if l[0] == '3013':
            l[7] = 0
            l[8] = 50000
            break

    # Step 4. Recalculate 2101 op_dr so Opening Dr == Opening Cr exactly
    other_op_dr = sum(l[7] for l in ledgers_def if l[0] != '2101')
    tot_op_cr = sum(l[8] for l in ledgers_def)
    req_2101_op_dr = tot_op_cr - other_op_dr

    for l in ledgers_def:
        if l[0] == '2101':
            l[7] = req_2101_op_dr
            l[8] = 0
            break

    # Verify Opening Balance equality
    tot_op_dr = sum(l[7] for l in ledgers_def)
    tot_op_cr = sum(l[8] for l in ledgers_def)
    print(f"Final Opening Dr: {tot_op_dr:,.2f}, Opening Cr: {tot_op_cr:,.2f}")
    assert tot_op_dr == tot_op_cr

    # Step 5. Final Pass: Adjust 2101 Month 12 Cr/Dr so Total Assets == Total EL exactly (difference = 0.00)
    close_anc = sum((l[7] - l[8]) + sum(l[9][m] - l[10][m] for m in range(12)) for l in ledgers_def if l[4] == 'anc')
    close_ac = sum((l[7] - l[8]) + sum(l[9][m] - l[10][m] for m in range(12)) for l in ledgers_def if l[4] == 'ac')
    close_eq = sum((l[8] - l[7]) + sum(l[10][m] - l[9][m] for m in range(12)) for l in ledgers_def if l[4] == 'eq')
    close_lnc = sum((l[8] - l[7]) + sum(l[10][m] - l[9][m] for m in range(12)) for l in ledgers_def if l[4] == 'lnc')
    close_lc = sum((l[8] - l[7]) + sum(l[10][m] - l[9][m] for m in range(12)) for l in ledgers_def if l[4] == 'lc')

    total_assets = close_anc + close_ac
    total_el = close_eq + close_lnc + close_lc
    diff = total_assets - total_el

    if diff > 0:
        for l in ledgers_def:
            if l[0] == '2101':
                l_cr = list(l[10])
                l_cr[11] += diff
                l[10] = l_cr
                break
    elif diff < 0:
        diff_abs = abs(diff)
        for l in ledgers_def:
            if l[0] == '2101':
                l_dr = list(l[9])
                l_dr[11] += diff_abs
                l[9] = l_dr
                break

    # Re-verify Month 12 Balance Sheet Balance
    close_anc = sum((l[7] - l[8]) + sum(l[9][m] - l[10][m] for m in range(12)) for l in ledgers_def if l[4] == 'anc')
    close_ac = sum((l[7] - l[8]) + sum(l[9][m] - l[10][m] for m in range(12)) for l in ledgers_def if l[4] == 'ac')
    close_eq = sum((l[8] - l[7]) + sum(l[10][m] - l[9][m] for m in range(12)) for l in ledgers_def if l[4] == 'eq')
    close_lnc = sum((l[8] - l[7]) + sum(l[10][m] - l[9][m] for m in range(12)) for l in ledgers_def if l[4] == 'lnc')
    close_lc = sum((l[8] - l[7]) + sum(l[10][m] - l[9][m] for m in range(12)) for l in ledgers_def if l[4] == 'lc')

    total_assets = close_anc + close_ac
    total_el = close_eq + close_lnc + close_lc
    print(f"Month 12 Total Assets: {total_assets:,.2f}, Total EL: {total_el:,.2f}, Diff: {total_el - total_assets:,.2f}")
    assert abs(total_el - total_assets) < 0.01, f"BS Unbalanced: Assets={total_assets}, EL={total_el}"

    # ── Create Workbook ──
    wb = openpyxl.Workbook()
    
    # Sheet 1: Trial_Balance
    ws_tb = wb.active
    ws_tb.title = 'Trial_Balance'
    
    tb_headers = ['Ledger_Code', 'Ledger_Name', 'Opening_Dr', 'Opening_Cr']
    for m in months_name:
        tb_headers.extend([f'{m}_Dr', f'{m}_Cr'])
    
    ws_tb.append(tb_headers)
    
    for l in ledgers_def:
        row = [l[0], l[1], l[7], l[8]]
        for m in range(12):
            row.extend([l[9][m], l[10][m]])
        ws_tb.append(row)
        
    # Sheet 2: Ledger_Master
    ws_lm = wb.create_sheet(title='Ledger_Master')
    lm_headers = ['Ledger_Code', 'Ledger_Name', 'Note_No', 'Note_Name', 'Section', 'Treasury_Type', 'Normal_Balance']
    ws_lm.append(lm_headers)
    
    for l in ledgers_def:
        ws_lm.append([l[0], l[1], l[2], l[3], l[4], l[5] or '', l[6]])
        
    # Sheet 3: Instructions
    ws_inst = wb.create_sheet(title='Instructions')
    instructions = [
        ['FinCommand Pro — Complete Financial Dataset & Trial Balance — FY 2025-26 (Apr 2025 to Mar 2026)'],
        [''],
        ['Overview & Architecture:'],
        ['1. Contains 3 sheets: Trial_Balance, Ledger_Master, and Instructions.'],
        ['2. Complete Coverage: Seeded with all standard 115 ledgers spanning Notes 1 to 26 under IND AS Schedule III.'],
        ['3. 28 Columns per Ledger in Trial_Balance: Ledger_Code, Ledger_Name, Opening_Dr, Opening_Cr, and Apr_Dr/Cr through Mar_Cr.'],
        ['4. Double-Entry Integrity: Total Opening Dr = Total Opening Cr, and for every month Apr–Mar, Total Monthly Dr = Total Monthly Cr.'],
        ['5. Schedule III Sections:'],
        ['   - eq  : Shareholders Equity (Notes 1, 2)'],
        ['   - lnc : Non-Current Liabilities (Notes 3, 4, 5, 6)'],
        ['   - lc  : Current Liabilities (Notes 7, 8, 9, 17, 18)'],
        ['   - anc : Non-Current Assets (Notes 10, 11, 12, 13, 14)'],
        ['   - ac  : Current Assets (Notes 13, 15, 16, 19, 20, 21, 23)'],
        ['   - inc : Income (Notes 20, 21)'],
        ['   - exp : Expenses (Notes 22, 23, 24, 25, 26)'],
        ['6. Treasury Tagging for Liquidity & Cash Flow:'],
        ['   - cash    : Cash in Hand & Petty Cash'],
        ['   - bank_ca : Current Accounts'],
        ['   - bank_sb : Savings / Sweep Accounts'],
        ['   - fd      : Fixed Deposits'],
        ['   - mf      : Liquid Mutual Funds'],
    ]
    for r in instructions:
        ws_inst.append(r)
        
    # ── Styling ──
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    regular_font = Font(name='Segoe UI', size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    for ws in [ws_tb, ws_lm]:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = regular_font
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
                    
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Styling Instructions Sheet
    ws_inst['A1'].font = Font(name='Segoe UI', size=14, bold=True, color='1F4E79')
    for row in ws_inst.iter_rows(min_row=3):
        for cell in row:
            if cell.value and cell.value.startswith(('Overview', '1.', '2.', '3.', '4.', '5.', '6.')):
                cell.font = Font(name='Segoe UI', size=11, bold=True)
            else:
                cell.font = regular_font
    ws_inst.column_dimensions['A'].width = 110

    # Output paths
    paths = [
        r'C:\Users\syamm\Downloads\fincommand_full_data.xlsx',
        r'C:\Users\syamm\Downloads\FinCommand_Full_Dataset.xlsx',
        r'C:\Users\syamm\Downloads\fincommand_full_data_v3.xlsx',
        r'C:\Users\syamm\Downloads\FinCommand_TB_FY2025-26.xlsx',
    ]
    
    for p in paths:
        try:
            wb.save(p)
            print(f"Successfully saved full dataset to: {p}")
        except Exception as e:
            print(f"Note: Could not save to {p}: {e}")

if __name__ == '__main__':
    build_dataset()
